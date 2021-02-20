#!/usr/bin/env python
# -*- encoding: utf-8 -*-
'''
@File    :   douban.py
@Time    :   2021/02/18
@Author  :   HDUZN
@Version :   1.0
@Contact :   hduzn@vip.qq.com
@License :   (C)Copyright 2021-2022
@Desc    :   将Douban 读过的书的记录保存到douban.xlsx 文件和数据库中
             (注：豆瓣标记后，没写标签、没评分/星 会报错，没写评论没关系。)
'''

# here put the import lib
import douban_config
import z_db
import time
import selenium
from selenium import webdriver
import openpyxl
import os

# 获取最大页数和每一页的模板链接
def get_max_page_num(wd):
    num_element = wd.find_element_by_css_selector('.paginator a:nth-last-of-type(1)')
    max_page_num = int(num_element.text)
    site = num_element.get_attribute('href')
    num_site_list = [max_page_num, site]
    return num_site_list

# 获取所有页数的链接List
def get_page_site_list(max_page_num, model_site):
    # model_site: https://book.douban.com/people/[id]/collect?start=285&sort=time&rating=all&filter=all&mode=grid
    model_site_split1 = model_site.split('=', 1)
    site_part1 = model_site_split1[0] + '='
    # print(site_part1) # https://book.douban.com/people/[id]/collect?start=
    model_site_split2 = model_site_split1[1].split('&', 1)
    # site_part2_num = model_site_split2[0]
    # print(site_part2_num) # 285
    site_part3 = '&' + model_site_split2[1]
    # print(site_part3) # &sort=time&rating=all&filter=all&mode=grid
    # page_site = site_part1 + site_part2_num + site_part3

    site_list = []
    for i in range(max_page_num):
        site_part2_num = str(i * 15)
        page_site = site_part1 + site_part2_num + site_part3
        site_list.append(page_site)
    return site_list

# 获取读过的书的页面 wd
def get_readed_wd(site):
    driver_path = douban_config.driver_path
    wd = webdriver.Chrome(driver_path)
    wd.implicitly_wait(10)

    wd.get(site)
    time.sleep(2)
    # 登录
    wd.find_element_by_class_name('top-nav-info').click()
    wd.find_element_by_class_name('account-tab-account').click()
    account = wd.find_element_by_id('username')
    pwd = wd.find_element_by_id('password')
    login = wd.find_element_by_css_selector('.account-form-field-submit a')
    account.clear()
    pwd.clear()
    account.send_keys(douban_config.douban_id)
    pwd.send_keys(douban_config.douban_pwd)
    login.click()
    time.sleep(5)

    # 打开个人主页
    wd.find_element_by_css_selector('.nav-user-account a').click()
    wd.find_element_by_css_selector('.more-items tr a').click()
    time.sleep(1)

    # 读过（我这里只有想读、读过的记录，没有在读），如果有【在读】选项，type括号中的数字为(3)
    readed = wd.find_element_by_css_selector('#book h2 .pl a:nth-of-type(2)')
    readed_site = readed.get_attribute('href')
    readed.click()

    # switch to new window 切换到新的窗口
    for handle in wd.window_handles:
        wd.switch_to.window(handle)
        if(wd.current_url == readed_site):
            break

    time.sleep(1)
    return wd

# 获取一页的内容，一共15项
def get_one_page(wd, page_site):
    try:
        wd.get(page_site)
        time.sleep(2)
    except selenium.common.exceptions.TimeoutException:
        time.sleep(5)
        wd.get(page_site)

    pic_list = []
    pic_values = wd.find_elements_by_css_selector('.interest-list .subject-item .pic a img')
    for value in pic_values:
        pic_site = value.get_attribute('src')
        pic_list.append(pic_site)
    # print(pic_list)

    book_site_list = []
    book_name_list = []
    book_values = wd.find_elements_by_css_selector('.interest-list .subject-item .info h2 a')
    for value in book_values:
        book_site = value.get_attribute('href')
        book_site_list.append(book_site)
        book_name = value.text
        book_name_list.append(book_name)
    # print(book_site_list)
    # print(book_name_list)

    author_list = []
    author_info = wd.find_elements_by_css_selector('.interest-list .subject-item .info .pub')
    for value in author_info:
        author_list.append(value.text)
    # print(author_list)

    tags_list = []
    tag_values = wd.find_elements_by_css_selector('.interest-list .subject-item .info .tags')
    for value in tag_values:
        tags_list.append(value.text.split(':')[1].strip())
    # print(tags_list)

    date_list = []
    date_values = wd.find_elements_by_css_selector('.interest-list .subject-item .info .date')
    for value in date_values:
        date_list.append(value.text.split()[0])
    # print(date_list)

    comment_list = []
    comment_values = wd.find_elements_by_css_selector('.interest-list .subject-item .info .comment')
    for value in comment_values:
        comment_list.append(value.text)
    # print(comment_list)

    rating_list = []
    rating_values = wd.find_elements_by_css_selector('.interest-list .subject-item .info .short-note div span:nth-of-type(1)')
    for value in rating_values:
        rating_list.append(douban_config.rating_dict[value.get_attribute('class')[6]])
    # print(rating_list)

    num = len(book_name_list)
    data_list = []
    for i in range(num):
        data_list.append([book_name_list[i], book_site_list[i], author_list[i], tags_list[i], date_list[i], comment_list[i], rating_list[i], pic_list[i]])
    return data_list

# 把记录数据List 全部存入数据库
def insert_into_db(db_file, table_name, data_list):
    columns = ['name', 'site', 'author', 'tags', 'date', 'comments', 'rating', 'pic']

    sql = z_db.get_insert_sql_by_colum_names(table_name, columns)
    # print(sql) # insert into books (name, site, author, tags, date, comments, rating, pic) values (?, ?, ?, ?, ?, ?, ?, ?)
    z_db.insert_into_db(db_file, table_name, sql, data_list)

# 把记录数据List 全部存入Excel表
def write_to_excel(ex_file, sheet_name, data_list):
    if(os.path.exists(ex_file)):
        book = openpyxl.load_workbook(ex_file)
        if(book[sheet_name]):
            sheet = book[sheet_name]
        else:
            sheet = book.create_sheet(sheet_name)
    else:
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = sheet_name

    for data_line in data_list:
        sheet.append(data_line)
    book.save(ex_file)

# main
def main():
    site = douban_config.books_site
    wd = get_readed_wd(site)

    num_site_list = get_max_page_num(wd)
    max_page_num = num_site_list[0]
    print(max_page_num) # 20
    model_site = num_site_list[1]
    # print(model_site) # https://book.douban.com/people/[id]/collect?start=285&sort=time&rating=all&filter=all&mode=grid
    
    page_site_list = get_page_site_list(max_page_num, model_site)

    all_page_data_list = []
    for page_site in page_site_list:
        # print(page_site)
        print('------------num: ', page_site_list.index(page_site)+1)
        one_page_data_list = get_one_page(wd, page_site)
        all_page_data_list = all_page_data_list + one_page_data_list
    # print(all_page_data_list)
    wd.quit()

    all_page_data_list.reverse() # 倒序

    db_file = douban_config.db_file
    table_name = douban_config.books_table_name
    insert_into_db(db_file, table_name, all_page_data_list)
    time.sleep(3)

    ex_file = douban_config.ex_file
    sheet_name = douban_config.book_sheet_name
    write_to_excel(ex_file, sheet_name, all_page_data_list)

    print('--------------------- Douban 已读书目 记录保存成功！')

main()
# z_db.delete_table(douban_config.db_file, douban_config.books_table_name) # 清空表